"""
media-monitor — portable keyword scan of configured sites.
Configs and reports live next to the executable (USB-friendly).
"""

from __future__ import annotations

import getpass
import json
import os
import re
import shutil
import subprocess
import sys
import threading
import time
import warnings
from concurrent.futures import CancelledError as FuturesCancelledError
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from email.utils import parsedate_to_datetime
from pathlib import Path
from typing import Callable, Iterable, TextIO
from urllib.parse import quote, urljoin, urlparse

import certifi
import requests
import urllib3
from bs4 import BeautifulSoup, XMLParsedAsHTMLWarning
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
warnings.filterwarnings("ignore", category=XMLParsedAsHTMLWarning)

VERSION = "3.3.1"

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/122.0.0.0 Safari/537.36"
)
# (connect timeout, read timeout) — fail faster on dead/hanging hosts
REQUEST_TIMEOUT = (5, 20)
MAX_PAGE_CHARS = 500_000
LOG_NAME = "media-monitor_log.txt"
LOG_MAX_BYTES = 10 * 1024 * 1024  # 10 MB
DEFAULT_AUTH_TIMEOUT = 180

# Hosts that already failed cert verification this run — skip verify next time.
_INSECURE_SSL_HOSTS: set[str] = set()
_SSL_LOCK = threading.Lock()
_PRINT_LOCK = threading.RLock()
# Live HTTP sessions — closed on Stop to unblock workers stuck in network I/O.
_ACTIVE_SESSIONS: set[requests.Session] = set()
_SESSIONS_LOCK = threading.Lock()
# Path to certifi + bundled Минцифры/extra CAs; set in run().
_CA_BUNDLE: str | bool = True
# When False, only progress/HIT/summary go to console (errors suppressed).
_LOG_VERBOSE = False
# Optional UI/log sink: receives each printed line (including trailing \n).
_LOG_CALLBACK: Callable[[str], None] | None = None
# Optional detailed UI sink (full log tab) — always receives detail lines.
_DETAIL_CALLBACK: Callable[[str], None] | None = None
# Optional live progress sink for GUI status panel (dict snapshot).
_PROGRESS_CALLBACK: Callable[[dict], None] | None = None
# Cooperative cancel for GUI Stop button.
_CANCEL = threading.Event()


class AuthRequiredError(Exception):
    """Page requires authentication that was not provided or failed."""


class CancelledError(Exception):
    """Scan was cancelled by the user."""


def set_log_callback(callback: Callable[[str], None] | None) -> None:
    global _LOG_CALLBACK
    _LOG_CALLBACK = callback


def set_detail_callback(callback: Callable[[str], None] | None) -> None:
    """Full-detail UI sink (always on for the «Полный лог» tab)."""
    global _DETAIL_CALLBACK
    _DETAIL_CALLBACK = callback


def set_progress_callback(callback: Callable[[dict], None] | None) -> None:
    global _PROGRESS_CALLBACK
    _PROGRESS_CALLBACK = callback


def register_session(session: requests.Session) -> None:
    with _SESSIONS_LOCK:
        _ACTIVE_SESSIONS.add(session)


def unregister_session(session: requests.Session) -> None:
    with _SESSIONS_LOCK:
        _ACTIVE_SESSIONS.discard(session)


def _close_all_sessions() -> int:
    """Force-close live sessions so blocked GETs fail quickly on Stop."""
    with _SESSIONS_LOCK:
        sessions = list(_ACTIVE_SESSIONS)
    closed = 0
    for session in sessions:
        try:
            session.close()
            closed += 1
        except Exception:
            pass
    return closed


def request_cancel() -> None:
    _CANCEL.set()
    n = _close_all_sessions()
    if n:
        log_print(
            f"Стоп: закрываю {n} сетевых соединений, "
            "чтобы не ждать таймауты…"
        )


def clear_cancel() -> None:
    _CANCEL.clear()
    with _SESSIONS_LOCK:
        _ACTIVE_SESSIONS.clear()


def is_cancelled() -> bool:
    return _CANCEL.is_set()


def log_print(*args: object, **kwargs: object) -> None:
    """Thread-safe print (whole line under one lock) + optional UI sinks."""
    ensure_stdio()
    sep = str(kwargs.get("sep", " "))
    end = str(kwargs.get("end", "\n"))
    line = sep.join(str(a) for a in args) + end
    with _PRINT_LOCK:
        try:
            print(*args, **kwargs)
        except Exception:
            pass
        cb = _LOG_CALLBACK
        if cb is not None:
            try:
                cb(line)
            except Exception:
                pass
        detail_cb = _DETAIL_CALLBACK
        if detail_cb is not None:
            try:
                detail_cb(line)
            except Exception:
                pass


def detail_print(*args: object, **kwargs: object) -> None:
    """
    Detailed progress (URL/search/errors).
    Always goes to the GUI «Полный лог» tab when a detail callback is set.
    Also printed to console/file when log_verbose=1.
    """
    ensure_stdio()
    sep = str(kwargs.get("sep", " "))
    end = str(kwargs.get("end", "\n"))
    line = sep.join(str(a) for a in args) + end
    with _PRINT_LOCK:
        if _LOG_VERBOSE:
            try:
                print(*args, **kwargs)
            except Exception:
                pass
        detail_cb = _DETAIL_CALLBACK
        if detail_cb is not None:
            try:
                detail_cb(line)
            except Exception:
                pass


def app_dir() -> Path:
    """Directory of the running exe (frozen) or this script (dev)."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def resource_dir() -> Path:
    """Packaged resources (PyInstaller _MEIPASS) or project root in dev."""
    meipass = getattr(sys, "_MEIPASS", None)
    if meipass:
        return Path(meipass)
    return Path(__file__).resolve().parent


def build_ca_bundle() -> str:
    """certifi roots + certs/*.pem (Минцифры, intermediates) baked into the build."""
    import tempfile

    parts = [Path(certifi.where()).read_text(encoding="utf-8")]
    certs_dir = resource_dir() / "certs"
    extra = 0
    if certs_dir.is_dir():
        for pem in sorted(certs_dir.glob("*.pem")):
            parts.append(pem.read_text(encoding="utf-8"))
            extra += 1
    out = Path(tempfile.gettempdir()) / "media-monitor-ca-bundle.pem"
    out.write_text("\n".join(parts), encoding="utf-8")
    detail_print(f"SSL CA bundle: certifi + {extra} extra CA(s)")
    return str(out)


class Tee:
    """Write the same text to console and an appendable log file."""

    def __init__(self, *streams: TextIO | None) -> None:
        self.streams = tuple(s for s in streams if s is not None)

    def write(self, data: str) -> int:
        with _PRINT_LOCK:
            for stream in self.streams:
                try:
                    stream.write(data)
                    stream.flush()
                except Exception:
                    pass
        return len(data)

    def flush(self) -> None:
        with _PRINT_LOCK:
            for stream in self.streams:
                try:
                    stream.flush()
                except Exception:
                    pass


class _NullWriter:
    """Stand-in when the process has no console (windowed exe)."""

    def write(self, data: str) -> int:
        return len(data)

    def flush(self) -> None:
        return None


def ensure_stdio() -> None:
    """Avoid crashes in --windowed builds where stdout/stderr are None."""
    if sys.stdout is None:
        sys.stdout = _NullWriter()  # type: ignore[assignment]
    if sys.stderr is None:
        sys.stderr = _NullWriter()  # type: ignore[assignment]

@dataclass(frozen=True)
class Settings:
    # Articles with publish date on/after this day ("не старше"). Empty = no min.
    article_date_not_older_than: date | None
    # Articles with publish date on/before this day ("дата ДО"). Empty = today.
    article_date_not_later_than: date
    auth_timeout_seconds: int
    # Max pages per site (0 = unlimited).
    max_scan_urls: int
    max_expand_links: int
    # True = verify TLS certs (with per-host fallback cache); False = never verify.
    ssl_verify: bool
    # Parallel site workers: 0 = one thread per site (all at once).
    site_workers: int
    # 1 = log every URL/errors; 0 = only status lines.
    log_verbose: bool
    # If > 0, search only within last N days from today (overrides both date fields).
    article_date_last_days: int
    # Comment unavailable sites in sites.txt with leading #.
    comment_unavailable_sites: bool


@dataclass(frozen=True)
class Hit:
    phrase: str
    url: str
    published_at: str
    title: str
    scanned_at: datetime


@dataclass
class SiteResult:
    origin: str
    label: str
    hits: list[Hit] = field(default_factory=list)
    pages_collected: int = 0
    pages_scanned: int = 0
    errors: int = 0
    skipped_auth: int = 0
    elapsed: float = 0.0
    fatal: str | None = None
    unavailable: bool = False


@dataclass
class RunProgress:
    """Live counters for UI / status lines (thread-safe)."""

    total_sites: int
    workers: int = 0
    _lock: threading.Lock = field(default_factory=threading.Lock)
    done: int = 0
    hits: int = 0
    pages: int = 0
    unavailable: int = 0
    cancelled_sites: int = 0
    active: set[str] = field(default_factory=set)

    def site_start(self, label: str) -> None:
        with self._lock:
            self.active.add(label)
        self._notify()

    def site_done(
        self,
        label: str,
        hit_count: int,
        *,
        pages: int = 0,
        unavailable: bool = False,
        cancelled: bool = False,
    ) -> tuple[int, int]:
        with self._lock:
            self.active.discard(label)
            self.done += 1
            self.hits += hit_count
            self.pages += pages
            if unavailable:
                self.unavailable += 1
            if cancelled:
                self.cancelled_sites += 1
            done = self.done
            hits = self.hits
        self._notify()
        return done, hits

    def snapshot(self) -> dict:
        with self._lock:
            names = sorted(self.active)
            return {
                "done": self.done,
                "total": self.total_sites,
                "hits": self.hits,
                "pages": self.pages,
                "unavailable": self.unavailable,
                "cancelled_sites": self.cancelled_sites,
                "active": len(self.active),
                "active_names": names[:6],
                "workers": self.workers,
                "cancelling": is_cancelled(),
            }

    def _notify(self) -> None:
        cb = _PROGRESS_CALLBACK
        if cb is None:
            return
        try:
            cb(self.snapshot())
        except Exception:
            pass


def ensure_local_config(root: Path, name: str) -> Path:
    """
    Use local config file.
    If missing, copy from *.example.txt once — never overwrite an existing file.
    """
    target = root / name
    if target.is_file():
        return target
    example = root / name.replace(".txt", ".example.txt")
    if example.is_file():
        shutil.copy2(example, target)
        print(f"Created {name} from {example.name} (edit it for your list).")
        return target
    raise FileNotFoundError(
        f"Missing {name} and {example.name}. "
        "Add the example file from the repo or create the config manually."
    )


def resolve_settings_path(root: Path) -> Path:
    """Only settings.txt. Migrate legacy setting.txt once if needed."""
    settings = root / "settings.txt"
    legacy = root / "setting.txt"
    if not settings.is_file() and legacy.is_file():
        shutil.copy2(legacy, settings)
        print("Migrated setting.txt -> settings.txt (use only settings.txt from now on).")
    return ensure_local_config(root, "settings.txt")


def load_lines(path: Path) -> list[str]:
    if not path.is_file():
        raise FileNotFoundError(f"Config not found: {path.name}")
    lines: list[str] = []
    for raw in path.read_text(encoding="utf-8-sig").splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        lines.append(line)
    return lines


def load_settings(path: Path) -> Settings:
    raw: dict[str, str] = {}
    for line in path.read_text(encoding="utf-8-sig").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        raw[key.strip().lower()] = value.strip()

    min_date: date | None = None
    older_raw = raw.get("article_date_not_older_than", "")
    if older_raw:
        min_date = date.fromisoformat(older_raw)

    # "дата ДО": empty means today
    later_raw = raw.get("article_date_not_later_than", "")
    max_date = date.fromisoformat(later_raw) if later_raw else date.today()

    timeout = DEFAULT_AUTH_TIMEOUT
    timeout_raw = raw.get("auth_timeout_seconds", "")
    if timeout_raw:
        timeout = max(1, int(timeout_raw))

    max_scan = 0
    max_scan_raw = raw.get("max_scan_urls", "")
    if max_scan_raw:
        max_scan = max(0, int(max_scan_raw))

    max_expand = 120
    max_expand_raw = raw.get("max_expand_links", "")
    if max_expand_raw:
        max_expand = max(1, int(max_expand_raw))

    ssl_verify = True
    ssl_raw = raw.get("ssl_verify", "1").strip().lower()
    if ssl_raw in {"0", "false", "no", "off", "нет"}:
        ssl_verify = False

    site_workers = 0
    workers_raw = raw.get("site_workers", "")
    if workers_raw:
        site_workers = max(0, int(workers_raw))

    log_verbose = False
    verbose_raw = raw.get("log_verbose", "0").strip().lower()
    if verbose_raw in {"1", "true", "yes", "on", "да"}:
        log_verbose = True

    article_date_last_days = 0
    recent_raw = raw.get("article_date_last_days", "")
    if recent_raw:
        article_date_last_days = max(0, int(recent_raw))

    comment_unavailable = True
    comment_raw = raw.get("comment_unavailable_sites", "1").strip().lower()
    if comment_raw in {"0", "false", "no", "off", "нет"}:
        comment_unavailable = False

    return Settings(
        article_date_not_older_than=min_date,
        article_date_not_later_than=max_date,
        auth_timeout_seconds=timeout,
        max_scan_urls=max_scan,
        max_expand_links=max_expand,
        ssl_verify=ssl_verify,
        site_workers=site_workers,
        log_verbose=log_verbose,
        article_date_last_days=article_date_last_days,
        comment_unavailable_sites=comment_unavailable,
    )


def ensure_settings_option(
    path: Path,
    key: str,
    default: str,
    comments: list[str],
) -> bool:
    """
    If settings.txt has no key=… yet, append comments + key=default.
    Never changes an existing value.
    """
    if not path.is_file():
        return False
    text = path.read_text(encoding="utf-8-sig")
    for line in text.splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or "=" not in stripped:
            continue
        if stripped.split("=", 1)[0].strip().lower() == key.lower():
            return False
    comment_block = "\n".join(f"# {c}" if c else "#" for c in comments)
    addition = f"\n{comment_block}\n{key}={default}\n"
    if text and not text.endswith("\n"):
        addition = "\n" + addition
    path.write_text(text + addition, encoding="utf-8")
    log_print(f"settings.txt: добавлен параметр {key}={default}")
    return True


def sync_settings_file(path: Path) -> None:
    """Append newly introduced settings keys without touching existing values."""
    ensure_settings_option(
        path,
        "site_workers",
        "0",
        [
            "Сколько сайтов сканировать одновременно (каждый сайт — свой поток).",
            "Внутри сайта запросы всегда идут по одному.",
            "0 = все сайты сразу (рекомендуется).",
            "Например 8 — не больше 8 сайтов параллельно.",
        ],
    )
    ensure_settings_option(
        path,
        "log_verbose",
        "0",
        [
            "Подробный лог: 0 = только понятный статус (без URL/HIT/ошибок), 1 = URL/SSL/ошибки.",
        ],
    )
    ensure_settings_option(
        path,
        "article_date_last_days",
        "0",
        [
            "За сколько последних дней от сегодня искать статьи.",
            "0 = выключено (используются article_date_not_older_than/not_later_than).",
            "Если > 0 — этот параметр перекрывает оба фильтра дат.",
        ],
    )
    ensure_settings_option(
        path,
        "comment_unavailable_sites",
        "1",
        [
            "Если сайт полностью недоступен — закомментировать его строку в sites.txt (#).",
            "1 = да, 0 = нет. В Excel недоступные сайты всё равно пишутся в конце отчёта (красным).",
        ],
    )


def site_host_key(url: str) -> str:
    """Normalize host for matching sites.txt lines (lowercase, without www.)."""
    host = urlparse(normalize_url(url)).netloc.lower()
    if host.startswith("www."):
        host = host[4:]
    return host


def comment_unavailable_in_sites(sites_path: Path, unavailable_hosts: set[str]) -> int:
    """Comment out matching active lines in sites.txt. Returns how many lines changed."""
    if not unavailable_hosts or not sites_path.is_file():
        return 0
    text = sites_path.read_text(encoding="utf-8-sig")
    lines = text.splitlines()
    stamp = date.today().isoformat()
    changed = 0
    new_lines: list[str] = []
    for line in lines:
        stripped = line.strip()
        if not stripped or stripped.startswith("#"):
            new_lines.append(line)
            continue
        host = site_host_key(stripped)
        if host in unavailable_hosts:
            new_lines.append(f"# {stripped}  # недоступен {stamp}")
            changed += 1
        else:
            new_lines.append(line)
    if changed:
        ending = "\n" if text.endswith("\n") or not text else ""
        sites_path.write_text("\n".join(new_lines) + ending, encoding="utf-8")
    return changed


def normalize_url(url: str) -> str:
    url = url.strip()
    if not url.startswith(("http://", "https://")):
        url = "https://" + url
    return url


def normalize_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def timed_input(prompt: str, timeout_sec: float) -> str | None:
    """Read a line from stdin; return None on timeout."""
    if timeout_sec <= 0:
        print(f"{prompt}(timeout)")
        return None

    result: list[str | None] = [None]
    done = threading.Event()

    def worker() -> None:
        try:
            # Print prompt to real console so password flow stays usable under Tee.
            sys.__stdout__.write(prompt)
            sys.__stdout__.flush()
            result[0] = sys.__stdin__.readline()
        except Exception:
            result[0] = None
        finally:
            done.set()

    thread = threading.Thread(target=worker, daemon=True)
    thread.start()
    if not done.wait(timeout_sec):
        print("(timeout)")
        return None
    value = result[0]
    if value is None:
        return None
    return value.rstrip("\r\n")


def timed_password(prompt: str, timeout_sec: float) -> str | None:
    """Read password without echo; return None on timeout."""
    if timeout_sec <= 0:
        print(f"{prompt}(timeout)")
        return None

    result: list[str | None] = [None]
    done = threading.Event()

    def worker() -> None:
        try:
            result[0] = getpass.getpass(prompt)
        except Exception:
            result[0] = None
        finally:
            done.set()

    thread = threading.Thread(target=worker, daemon=True)
    thread.start()
    if not done.wait(timeout_sec):
        print("(timeout)")
        return None
    return result[0]


def prompt_credentials(timeout_sec: int) -> tuple[str, str] | None:
    """
    Ask for login/password for sites that need auth.
    Wait up to timeout_sec total; on timeout or empty login — skip auth.
    """
    print()
    print(
        "Авторизация для соцсетей (необязательно). "
        f"На ввод логина и пароля — до {timeout_sec} сек. "
        "Пустой логин или таймаут = пропуск."
    )
    deadline = datetime.now().timestamp() + timeout_sec

    login = timed_input("Логин: ", deadline - datetime.now().timestamp())
    if login is None:
        print("Авторизация пропущена (таймаут логина).")
        return None
    login = login.strip()
    if not login:
        print("Авторизация пропущена (пустой логин).")
        return None

    password = timed_password("Пароль: ", deadline - datetime.now().timestamp())
    if password is None:
        print("Авторизация пропущена (таймаут пароля).")
        return None

    print("Учётные данные приняты (HTTP Basic Auth для 401/403).")
    return login, password


def fetch_html(
    url: str,
    session: requests.Session,
    auth: tuple[str, str] | None = None,
    ssl_verify: bool = True,
) -> str:
    host = urlparse(url).netloc.lower()

    def _get(verify: bool | str, use_auth: bool) -> requests.Response:
        return session.get(
            url,
            timeout=REQUEST_TIMEOUT,
            allow_redirects=True,
            auth=auth if use_auth else None,
            verify=verify,
        )

    with _SSL_LOCK:
        insecure = host in _INSECURE_SSL_HOSTS
    if ssl_verify and not insecure:
        verify: bool | str = _CA_BUNDLE
    else:
        verify = False
    try:
        resp = _get(verify=verify, use_auth=False)
    except requests.exceptions.SSLError:
        first = False
        with _SSL_LOCK:
            if host not in _INSECURE_SSL_HOSTS:
                _INSECURE_SSL_HOSTS.add(host)
                first = True
        if first:
            detail_print(
                f"  [ssl] у {host} битый сертификат — дальше для этого сайта "
                "без проверки SSL (быстрее)."
            )
        resp = _get(verify=False, use_auth=False)

    if resp.status_code in (401, 403) and auth:
        try:
            with _SSL_LOCK:
                insecure = host in _INSECURE_SSL_HOSTS
            auth_verify: bool | str = (
                _CA_BUNDLE if ssl_verify and not insecure else False
            )
            resp = _get(verify=auth_verify, use_auth=True)
        except requests.exceptions.SSLError:
            with _SSL_LOCK:
                _INSECURE_SSL_HOSTS.add(host)
            resp = _get(verify=False, use_auth=True)
    if resp.status_code in (401, 403):
        raise AuthRequiredError(f"HTTP {resp.status_code} auth required/failed")
    resp.raise_for_status()
    resp.encoding = resp.apparent_encoding or resp.encoding or "utf-8"
    return resp.text


def extract_title(soup: BeautifulSoup) -> str:
    og = soup.find("meta", property="og:title")
    if og and og.get("content"):
        return normalize_spaces(str(og["content"]))

    if soup.title and soup.title.string:
        return normalize_spaces(soup.title.string)

    h1 = soup.find("h1")
    if h1:
        return normalize_spaces(h1.get_text(" ", strip=True))

    return ""


def parse_date_value(raw: str) -> date | None:
    value = raw.strip()
    if not value:
        return None

    # Relative Russian dates (common in feeds / some article headers).
    low = value.lower()
    if re.match(r"^сегодня\b", low):
        return date.today()
    if re.match(r"^вчера\b", low):
        return date.today() - timedelta(days=1)
    if re.match(r"^позавчера\b", low):
        return date.today() - timedelta(days=2)

    # ISO / HTML5 datetime
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00")).date()
    except ValueError:
        pass

    head = re.split(r"[T\s,]", value, maxsplit=1)[0]
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(head, fmt).date()
        except ValueError:
            continue

    # Dates written in text, e.g. "5 августа 2026", "27 декабря 2024 года, 20:27".
    months_ru = {
        "января": 1,
        "январь": 1,
        "янв.": 1,
        "янв": 1,
        "февраля": 2,
        "февраль": 2,
        "февр.": 2,
        "фев": 2,
        "марта": 3,
        "март": 3,
        "март.": 3,
        "мар": 3,
        "апреля": 4,
        "апрель": 4,
        "апр.": 4,
        "апр": 4,
        "мая": 5,
        "май": 5,
        "май.": 5,
        "июня": 6,
        "июнь": 6,
        "июн.": 6,
        "июн": 6,
        "июля": 7,
        "июль": 7,
        "июл.": 7,
        "июл": 7,
        "августа": 8,
        "август": 8,
        "авг.": 8,
        "авг": 8,
        "сентября": 9,
        "сентябрь": 9,
        "сен.": 9,
        "сен": 9,
        "сент": 9,
        "октября": 10,
        "октябрь": 10,
        "окт.": 10,
        "окт": 10,
        "ноября": 11,
        "ноябрь": 11,
        "ноя.": 11,
        "ноя": 11,
        "декабря": 12,
        "декабрь": 12,
        "дек.": 12,
        "дек": 12,
    }
    months_en = {
        "jan": 1,
        "january": 1,
        "feb": 2,
        "february": 2,
        "mar": 3,
        "march": 3,
        "apr": 4,
        "april": 4,
        "may": 5,
        "jun": 6,
        "june": 6,
        "jul": 7,
        "july": 7,
        "aug": 8,
        "august": 8,
        "sep": 9,
        "sept": 9,
        "september": 9,
        "oct": 10,
        "october": 10,
        "nov": 11,
        "november": 11,
        "dec": 12,
        "december": 12,
    }

    m = re.search(
        r"(?i)(\d{1,2})\s*([a-zа-яё\.]+)\s*(\d{4})(?:\s*г(?:ода|\.)?)?",
        value,
    )
    if m:
        day = int(m.group(1))
        mon_raw = m.group(2).strip().lower().rstrip(".")
        year = int(m.group(3))
        mon = months_ru.get(mon_raw) or months_ru.get(mon_raw + ".") or months_en.get(mon_raw)
        if mon:
            try:
                return date(year, mon, day)
            except ValueError:
                return None

    # RFC 2822
    try:
        return parsedate_to_datetime(value).date()
    except (TypeError, ValueError, IndexError):
        return None


def _looks_like_relative_date_text(raw: str) -> bool:
    t = raw.strip().lower()
    if re.match(r"^(сегодня|вчера|позавчера)\b", t):
        return True
    if re.match(r"^\d+\s*(час|часа|часов|мин|минут)\b", t):
        return True
    return False



ARTICLE_JSONLD_TYPES = {
    "newsarticle",
    "article",
    "blogposting",
    "reportagenewsarticle",
    "analysisnewsarticle",
    "opinionnewsarticle",
    "reviewnewsarticle",
}


def _jsonld_types(obj: dict) -> set[str]:
    raw = obj.get("@type")
    if isinstance(raw, str):
        return {raw.lower()}
    if isinstance(raw, list):
        return {str(x).lower() for x in raw}
    return set()


def extract_date_from_jsonld(soup: BeautifulSoup) -> date | None:
    """Prefer datePublished only from Article/NewsArticle JSON-LD nodes."""

    def walk(obj: object) -> date | None:
        if isinstance(obj, dict):
            types = _jsonld_types(obj)
            if types & ARTICLE_JSONLD_TYPES:
                for key in ("datePublished", "dateCreated", "uploadDate"):
                    value = obj.get(key)
                    if isinstance(value, str):
                        parsed = parse_date_value(value)
                        if parsed:
                            return parsed
            for value in obj.values():
                nested = walk(value)
                if nested:
                    return nested
        elif isinstance(obj, list):
            for item in obj:
                nested = walk(item)
                if nested:
                    return nested
        return None

    for script in soup.find_all("script", attrs={"type": lambda x: x and "ld+json" in x}):
        raw = script.string or script.get_text(strip=True)
        if not raw:
            continue
        try:
            data = json.loads(raw)
        except Exception:
            continue
        parsed = walk(data)
        if parsed:
            return parsed
    return None


def page_looks_like_article(soup: BeautifulSoup, url: str) -> bool:
    """
    Company cards / listings often have WordPress article:* meta but are not news.
    """
    path = urlparse(url).path.lower()
    non_article_prefixes = (
        "/kompanii/",
        "/company/",
        "/companies/",
        "/category/",
        "/tag/",
        "/author/",
        "/page/",
        "/about/",
        "/contact",
        "/reklama",
        "/advert",
    )
    if any(path.startswith(p) or f"{p.rstrip('/')}/" in path + "/" for p in non_article_prefixes):
        # /category/ and /kompanii/ are not articles to report
        if path.startswith(("/kompanii/", "/company/", "/companies/", "/category/", "/tag/", "/author/")):
            return False

    og_type = soup.find("meta", property="og:type")
    if og_type and str(og_type.get("content", "")).lower() in {"website", "profile", "business.business"}:
        # Still allow if JSON-LD says NewsArticle
        pass

    for script in soup.find_all("script", attrs={"type": lambda x: x and "ld+json" in x}):
        raw = script.string or script.get_text(strip=True)
        if not raw:
            continue
        try:
            data = json.loads(raw)
        except Exception:
            continue

        def has_article(obj: object) -> bool:
            if isinstance(obj, dict):
                if _jsonld_types(obj) & ARTICLE_JSONLD_TYPES:
                    return True
                return any(has_article(v) for v in obj.values())
            if isinstance(obj, list):
                return any(has_article(x) for x in obj)
            return False

        if has_article(data):
            return True

    if og_type and str(og_type.get("content", "")).lower() in {"article", "news"}:
        return True

    # Path heuristic: /YYYY/MM/slug, /YYYYMMDD-..., /news/
    if re.search(r"/\d{4}/\d{2}/", path):
        return True
    if re.search(r"/\d{8}(?:-\d+)?/?", path):
        return True
    if "/news/" in path or path.rstrip("/").endswith("/news"):
        return True

    # Explicit non-article sections
    if path.startswith(("/kompanii/", "/company/", "/companies/")):
        return False

    # Default: treat as candidate article if it has an article body node
    for sel in (
        "article .entry-content",
        ".entry-content",
        ".post-content",
        ".td-post-content",
        ".detailed-page",
        ".news-detail",
        "article",
    ):
        node = soup.select_one(sel)
        if node and len(node.get_text(" ", strip=True)) >= 200:
            # But reject short company cards wrapped in <article>
            if path.startswith(("/kompanii/", "/company/", "/companies/", "/category/")):
                return False
            return True
    return False


def extract_date_from_url(url: str) -> date | None:
    path = urlparse(url).path
    m = re.search(r"/(\d{4})(\d{2})(\d{2})(?:-|/|$)", path)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass
    m = re.search(r"/(\d{4})/(\d{2})/(\d{2})(?:/|$)", path)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass
    return None


def extract_date_from_visible(soup: BeautifulSoup) -> date | None:
    """
    Pull a publish date from visible page chrome.
    Prefer article-specific blocks (e.g. dostup1 .subtitle__date) over feed
    widgets that often say only «сегодня, 18:27».
    """
    preferred = (
        ".subtitle__date",
        ".article-date",
        ".article__date",
        ".news-date",
        ".publication-date",
        ".pub-date",
        ".detailed-page__date",
        ".material-date",
        ".b-article__date",
        ".post-date",
        ".entry-date",
        "time[datetime]",
        "[itemprop='datePublished']",
        "[class*='subtitle__date']",
        "[class*='article__date']",
        "[class*='article-date']",
        "[class*='publish']",
    )
    broad = (
        "time",
        ".date",
        "[class*='date']",
        "[class*='Date']",
    )

    def _raw_from(el) -> str:
        return str(
            el.get("datetime")
            or el.get("content")
            or el.get_text(" ", strip=True)
            or ""
        ).strip()

    # 1) Specific article date blocks — absolute dates first.
    for sel in preferred:
        for el in soup.select(sel)[:15]:
            raw = _raw_from(el)
            if not raw:
                continue
            if _looks_like_relative_date_text(raw):
                continue
            parsed = parse_date_value(raw)
            if parsed:
                return parsed

    # 2) Broad selectors — only absolute dates (skip feed «сегодня»).
    for sel in broad:
        for el in soup.select(sel)[:40]:
            raw = _raw_from(el)
            if not raw or _looks_like_relative_date_text(raw):
                continue
            parsed = parse_date_value(raw)
            if parsed:
                return parsed

    # 3) Relative date only from preferred article chrome.
    for sel in preferred:
        for el in soup.select(sel)[:15]:
            raw = _raw_from(el)
            if not raw:
                continue
            parsed = parse_date_value(raw)
            if parsed:
                return parsed
    return None



def extract_published_date(soup: BeautifulSoup, url: str = "") -> date | None:
    # Skip only obvious non-articles (company cards / categories).
    if url and not page_looks_like_article(soup, url) and not is_probable_article_url(url):
        return None
    path = urlparse(url).path.lower() if url else ""
    if path.startswith(("/kompanii/", "/company/", "/companies/", "/category/", "/tag/")):
        return None

    jsonld_date = extract_date_from_jsonld(soup)
    if jsonld_date:
        return jsonld_date

    meta_candidates: list[str] = []
    for prop in ("article:published_time", "og:published_time"):
        tag = soup.find("meta", property=prop)
        if tag and tag.get("content"):
            meta_candidates.append(str(tag["content"]))

    for name in (
        "pubdate",
        "publish-date",
        "publication_date",
        "date",
        "DC.date",
        "dc.date",
        "sailthru.date",
        "MediatorArticlePublishDate",
    ):
        tag = soup.find("meta", attrs={"name": name})
        if tag and tag.get("content"):
            meta_candidates.append(str(tag["content"]))

    tag = soup.find(attrs={"itemprop": "datePublished"})
    if tag:
        content = tag.get("content") or tag.get("datetime") or tag.get_text(" ", strip=True)
        if content:
            meta_candidates.append(str(content))

    for time_tag in soup.find_all("time"):
        content = time_tag.get("datetime") or time_tag.get_text(" ", strip=True)
        if content:
            meta_candidates.append(str(content))

    for raw in meta_candidates:
        parsed = parse_date_value(raw)
        if parsed:
            return parsed

    visible = extract_date_from_visible(soup)
    if visible:
        return visible

    if url:
        return extract_date_from_url(url)
    return None


def extract_main_text(soup: BeautifulSoup) -> str:
    """
    Search phrases only in main content — not header/nav/footer/sidebar chrome
    (otherwise hits like 'Срочные новости – в нашем Telegram' leak into reports).
    """
    work = BeautifulSoup(str(soup), "lxml")
    for tag in work(
        ["script", "style", "noscript", "svg", "template", "nav", "header", "footer", "aside"]
    ):
        tag.decompose()

    junk_re = re.compile(
        r"(sidebar|widget|menu|nav|footer|header|share|social|telegram|subscribe|banner|advert|rekla)",
        re.I,
    )
    for tag in list(work.find_all(True)):
        if not getattr(tag, "attrs", None):
            continue
        cid = " ".join(tag.get("class") or [])
        tid = str(tag.get("id") or "")
        if junk_re.search(f"{cid} {tid}"):
            tag.decompose()

    for sel in (
        "article .entry-content",
        ".entry-content",
        ".post-content",
        ".td-post-content",
        ".article-body",
        ".post-body",
        "article",
        "main",
        "[role=main]",
        "#content",
    ):
        node = work.select_one(sel)
        if node:
            text = node.get_text(separator=" ", strip=True)
            if len(text) >= 80:
                return text[:MAX_PAGE_CHARS]

    text = work.get_text(separator=" ", strip=True)
    return text[:MAX_PAGE_CHARS]


NON_ARTICLE_PATH_PREFIXES = (
    "/kompanii/",
    "/company/",
    "/companies/",
    "/category/",
    "/tag/",
    "/author/",
    "/page/",
    "/wp-admin/",
    "/wp-login",
    "/about/",
    "/contact",
    "/reklama",
    "/advert",
    "/feed",
    "/search",
    "/login",
    "/register",
    "/news/hashtag/",
    "/news/popular/",
    "/news/articles/",
    "/news/longread/",
    "/news/main/",
    "/russia/news/",
)

# Bare section roots (listings, not articles)
NON_ARTICLE_EXACT_PATHS = {
    "/news/",
    "/society/",
    "/economics/",
    "/economy/",
    "/finance/",
    "/politics/",
    "/sport/",
    "/culture/",
    "/testy/",
}


def is_probable_article_url(url: str) -> bool:
    path = urlparse(url).path.lower().rstrip("/") + "/"
    if path == "/":
        return False
    if path in NON_ARTICLE_EXACT_PATHS:
        return False
    for prefix in NON_ARTICLE_PATH_PREFIXES:
        if path.startswith(prefix) or prefix.rstrip("/") == path.rstrip("/"):
            return False
    if re.search(r"/page/\d+/?", path):
        return False
    return True


def article_link_priority(url: str) -> int:
    """Lower = better candidate (prefer concrete news IDs over section hubs)."""
    path = urlparse(url).path.lower()
    if re.search(r"/news/\d+", path):
        return 0
    if re.search(r"/\d{8}(?:-\d+)?", path):
        return 1
    if re.search(r"/\d{4}/\d{2}/", path):
        return 2
    if re.search(r"/\d{6,}", path):
        return 3
    return 9


def extract_page(html: str, base_url: str) -> tuple[str, date | None, str, list[str]]:
    """Return (title, published_date, main_text, same-host article-like links)."""
    soup = BeautifulSoup(html, "lxml")
    title = extract_title(soup)
    published = extract_published_date(soup, base_url)
    text = extract_main_text(soup)

    host = urlparse(base_url).netloc.lower()
    links: list[str] = []
    seen: set[str] = set()
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if href.startswith(("#", "javascript:", "mailto:", "tel:")):
            continue
        absolute = urljoin(base_url, href)
        parsed = urlparse(absolute)
        if parsed.scheme not in ("http", "https"):
            continue
        if parsed.netloc.lower() != host:
            continue
        clean = absolute.split("#", 1)[0]
        if not is_probable_article_url(clean):
            continue
        if clean not in seen:
            seen.add(clean)
            links.append(clean)
    links.sort(key=article_link_priority)
    return title, published, text, links


def parse_search_line(line: str) -> tuple[str, str]:
    """
    words.txt line → (mode, value).

    mode "phrase": exact phrase (line in "..." or «...»);
    mode "words": all tokens must appear as whole words, any order.
    """
    raw = normalize_spaces(line)
    if len(raw) >= 2:
        pairs = (
            ('"', '"'),
            ("'", "'"),
            ("«", "»"),
            ("“", "”"),
            ("„", "“"),
        )
        for left, right in pairs:
            if raw.startswith(left) and raw.endswith(right):
                return "phrase", normalize_spaces(raw[len(left) : -len(right)])
    return "words", raw


def query_for_site_search(line: str) -> str:
    """Text to send to on-site search (without surrounding quotes)."""
    _mode, value = parse_search_line(line)
    return value


def phrase_present(text: str, phrase: str) -> bool:
    """
    Match rules for one words.txt line:
    - "карта Виза принимается в городе" — exact phrase (order fixed);
    - тройка карта принимается — all words present as whole words, any order.
    Case-insensitive; whitespace normalized.
    """
    haystack = normalize_spaces(text).casefold()
    mode, value = parse_search_line(phrase)
    if not value:
        return False
    needle = value.casefold()

    if mode == "phrase":
        return needle in haystack

    tokens = needle.split()
    if not tokens:
        return False
    for token in tokens:
        # Whole-word match (Latin + Cyrillic via Unicode \w).
        if not re.search(rf"(?<!\w){re.escape(token)}(?!\w)", haystack):
            return False
    return True


def effective_article_date_range(
    min_date: date | None,
    max_date: date,
    recent_days: int = 0,
    *,
    today: date | None = None,
) -> tuple[date | None, date]:
    """
    Compute the from/to window used by the article date filter.
    If recent_days > 0 — [today - recent_days; today], ignoring min/max.
    Otherwise — [min_date; max_date] (min_date may be None = no lower bound).
    """
    today = today or date.today()
    if recent_days > 0:
        lower = date.fromordinal(today.toordinal() - recent_days)
        return lower, today
    return min_date, max_date


def format_article_date_range_ru(
    min_date: date | None,
    max_date: date,
) -> str:
    """Human-readable «с … по …» for UI/logs."""
    if min_date is None:
        return f"с (без нижней границы) по {max_date.isoformat()}"
    return f"с {min_date.isoformat()} по {max_date.isoformat()}"


def passes_date_filter(
    published: date | None,
    min_date: date | None,
    max_date: date | None,
    recent_days: int = 0,
) -> bool:
    """
    Keep if min_date <= published <= max_date.
    If recent_days > 0, use [today - recent_days; today] and ignore min/max args.
    Unknown publish date — always keep (include in report).
    """
    if published is None:
        return True
    # max_date is always set by load_settings (defaults to today); keep Optional for callers.
    eff_min, eff_max = effective_article_date_range(
        min_date,
        max_date if max_date is not None else date.today(),
        recent_days,
    )
    if eff_min is not None and published < eff_min:
        return False
    if published > eff_max:
        return False
    return True


def normalize_exclude_url(url: str) -> str:
    url = normalize_url(url).strip()
    return url.split("#", 1)[0].rstrip("/")


def load_exclude_urls(path: Path) -> set[str]:
    return {normalize_exclude_url(line) for line in load_lines(path)}


def is_excluded_url(url: str, excluded: set[str]) -> bool:
    if not excluded:
        return False
    return normalize_exclude_url(url) in excluded


def scan_page(
    url: str,
    phrases: list[str],
    session: requests.Session,
    scanned_at: datetime,
    settings: Settings,
    auth: tuple[str, str] | None,
    excluded: set[str] | None = None,
) -> list[Hit]:
    if excluded and is_excluded_url(url, excluded):
        return []

    html = fetch_html(url, session, auth=auth, ssl_verify=settings.ssl_verify)
    soup = BeautifulSoup(html, "lxml")

    # Seeds may be categories; expanded company cards must not become report rows.
    if not is_probable_article_url(url) and not page_looks_like_article(soup, url):
        return []

    title = extract_title(soup)
    published = extract_published_date(soup, url)
    text = extract_main_text(soup)

    if not passes_date_filter(
        published,
        settings.article_date_not_older_than,
        settings.article_date_not_later_than,
        settings.article_date_last_days,
    ):
        return []

    published_str = published.isoformat() if published else ""
    hits: list[Hit] = []
    for phrase in phrases:
        if phrase_present(text, phrase):
            hits.append(
                Hit(
                    phrase=phrase,
                    url=url,
                    published_at=published_str,
                    title=title,
                    scanned_at=scanned_at,
                )
            )
    return hits


def site_search_urls(origin: str, phrase: str) -> list[str]:
    """Common on-site search endpoints (helps find older articles not on homepage)."""
    q = quote(query_for_site_search(phrase))
    base = origin.rstrip("/")
    return [
        # Yii/media sites (amurmedia.ru etc.)
        f"{base}/search/?SearchModel%5Bkeyword%5D={q}",
        f"{base}/search/?q={q}",
        f"{base}/?s={q}",
        f"{base}/search?query={q}",
        f"{base}/search?text={q}",
    ]


def group_seeds_by_origin(seed_urls: Iterable[str]) -> list[tuple[str, list[str]]]:
    """Group seed URLs by origin; preserve first-seen order of sites."""
    order: list[str] = []
    groups: dict[str, list[str]] = {}
    for seed in seed_urls:
        url = normalize_url(seed)
        parsed = urlparse(url)
        origin = f"{parsed.scheme}://{parsed.netloc}"
        if origin not in groups:
            groups[origin] = []
            order.append(origin)
        if url not in groups[origin]:
            groups[origin].append(url)
    return [(origin, groups[origin]) for origin in order]


def collect_urls_for_site(
    seed_urls: list[str],
    phrases: list[str],
    session: requests.Session,
    auth: tuple[str, str] | None,
    settings: Settings,
    site_label: str,
) -> list[str]:
    """
    Build scan list for one site (sequential requests):
    1) seed URLs
    2) site search for each phrase
    3) article links from listing/home pages
    max_scan_urls applies per site.
    """
    result: list[str] = []
    seen: set[str] = set()
    limit = settings.max_scan_urls  # 0 = unlimited (per site)
    expand_cap = settings.max_expand_links

    def full() -> bool:
        return limit > 0 and len(result) >= limit

    def add(url: str) -> None:
        if url in seen or full():
            return
        seen.add(url)
        result.append(url)

    def add_links_from(page_url: str) -> None:
        if is_cancelled():
            raise CancelledError()
        if full():
            return
        try:
            html = fetch_html(
                page_url, session, auth=auth, ssl_verify=settings.ssl_verify
            )
            _, _, _, links = extract_page(html, page_url)
            for link in links[:expand_cap]:
                add(link)
                if full():
                    return
        except CancelledError:
            raise
        except AuthRequiredError as exc:
            detail_print(f"  [{site_label}] [skip auth] cannot expand {page_url}: {exc}")
        except requests.RequestException as exc:
            if is_cancelled():
                raise CancelledError() from exc
            detail_print(f"  [{site_label}] [warn] cannot expand {page_url}: {exc}")

    for url in seed_urls:
        if is_cancelled():
            raise CancelledError()
        add(url)

    origin = f"{urlparse(seed_urls[0]).scheme}://{urlparse(seed_urls[0]).netloc}"
    for phrase in phrases:
        if is_cancelled():
            raise CancelledError()
        if full():
            break
        for search_url in site_search_urls(origin, phrase):
            if is_cancelled():
                raise CancelledError()
            if full():
                break
            detail_print(f"  [{site_label}] search: {search_url}")
            add_links_from(search_url)

    for url in seed_urls:
        if is_cancelled():
            raise CancelledError()
        if full():
            break
        path = urlparse(url).path.rstrip("/")
        if path.count("/") <= 2:
            add_links_from(url)

    if limit > 0 and len(result) >= limit:
        detail_print(f"  [{site_label}] [limit] max_scan_urls={limit} (на сайт)")
    return result


def process_site(
    index: int,
    total: int,
    origin: str,
    seeds: list[str],
    phrases: list[str],
    settings: Settings,
    auth: tuple[str, str] | None,
    excluded: set[str],
    scanned_at: datetime,
    progress: RunProgress,
) -> SiteResult:
    """One site = one worker; all HTTP for this site is sequential."""
    label = urlparse(origin).netloc or origin
    started = time.monotonic()
    if is_cancelled():
        result = SiteResult(origin=origin, label=label, fatal="cancelled")
        progress.site_done(label, 0, cancelled=True)
        return result

    progress.site_start(label)
    detail_print(f"→ старт {label} ({index}/{total})")

    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT, "Accept-Language": "ru,en;q=0.8"})
    register_session(session)

    result = SiteResult(origin=origin, label=label)
    try:
        if is_cancelled():
            raise CancelledError()

        probe_url = seeds[0] if seeds else origin
        detail_print(f"  [{label}] проверка доступности: {probe_url}")
        try:
            fetch_html(
                probe_url, session, auth=auth, ssl_verify=settings.ssl_verify
            )
            detail_print(f"  [{label}] сайт отвечает, собираю ссылки…")
        except AuthRequiredError:
            # Host responds — not "unavailable", auth handled later per page.
            detail_print(f"  [{label}] отвечает (нужна авторизация), собираю ссылки…")
            pass
        except requests.RequestException as exc:
            if is_cancelled():
                raise CancelledError() from exc
            result.errors += 1
            result.unavailable = True
            detail_print(f"  [{label}] [error] сайт недоступен: {exc}")

        if not result.unavailable:
            if is_cancelled():
                raise CancelledError()
            urls = collect_urls_for_site(
                seeds, phrases, session, auth, settings, label
            )
            if excluded:
                urls = [u for u in urls if not is_excluded_url(u, excluded)]
            result.pages_collected = len(urls)
            detail_print(f"  [{label}] к скану {len(urls)} стр.")

            for i, url in enumerate(urls, start=1):
                if is_cancelled():
                    raise CancelledError()
                detail_print(f"  [{label}] [{i}/{len(urls)}] {url}")
                try:
                    page_hits = scan_page(
                        url, phrases, session, scanned_at, settings, auth, excluded
                    )
                    result.pages_scanned += 1
                    if page_hits:
                        result.hits.extend(page_hits)
                        for hit in page_hits:
                            detail_print(
                                f"  [{label}] HIT «{hit.phrase}» {hit.url}"
                            )
                except AuthRequiredError as exc:
                    result.skipped_auth += 1
                    detail_print(f"  [{label}] [skip auth] {exc}")
                except requests.RequestException as exc:
                    if is_cancelled():
                        raise CancelledError() from exc
                    result.errors += 1
                    detail_print(f"  [{label}] [error] {exc}")

            # All collected pages failed to load (and none succeeded).
            if (
                not result.unavailable
                and result.pages_scanned == 0
                and result.errors > 0
                and result.skipped_auth == 0
            ):
                result.unavailable = True
    except CancelledError:
        result.fatal = "cancelled"
    except Exception as exc:  # noqa: BLE001 — site must not kill the pool
        if is_cancelled():
            result.fatal = "cancelled"
        else:
            result.fatal = str(exc)
            result.unavailable = True
            detail_print(f"[{index}/{total}] FAIL {label} — {exc}")
    finally:
        unregister_session(session)
        try:
            session.close()
        except Exception:
            pass

    result.elapsed = time.monotonic() - started
    cancelled = result.fatal == "cancelled"
    done, hit_total = progress.site_done(
        label,
        len(result.hits),
        pages=result.pages_scanned,
        unavailable=bool(result.unavailable and not cancelled),
        cancelled=cancelled,
    )
    dur = format_duration(result.elapsed)
    if cancelled:
        log_print(
            f"[{done}/{total}] остановлен  {label} "
            f"(успели {result.pages_scanned} стр., {len(result.hits)} находок, {dur})"
        )
    elif result.unavailable:
        log_print(f"[{done}/{total}] недоступен  {label} ({dur})")
    else:
        log_print(
            f"[{done}/{total}] готово  {label} — "
            f"{result.pages_scanned} стр., {len(result.hits)} находок ({dur}) "
            f"| всего находок: {hit_total}"
        )
    return result


def format_duration(seconds: float) -> str:
    total = int(round(seconds))
    h, rem = divmod(total, 3600)
    m, s = divmod(rem, 60)
    if h:
        return f"{h}ч {m}м {s}с"
    if m:
        return f"{m}м {s}с"
    return f"{s}с"


def try_open_file(path: Path) -> bool:
    """Open a file with the OS default app (Excel for .xlsx)."""
    try:
        if not path.is_file():
            return False
        if sys.platform == "win32":
            os.startfile(str(path))  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            subprocess.run(["open", str(path)], check=False)
        else:
            subprocess.run(["xdg-open", str(path)], check=False)
        return True
    except Exception as exc:  # noqa: BLE001
        log_print(f"Не удалось открыть файл: {exc}")
        return False


def rotate_log_if_needed(log_path: Path, max_bytes: int = LOG_MAX_BYTES) -> None:
    """Keep log file at most max_bytes by trimming oldest content."""
    if not log_path.is_file():
        return
    try:
        size = log_path.stat().st_size
    except OSError:
        return
    if size <= max_bytes:
        return
    try:
        data = log_path.read_bytes()
        # Keep the newest half of the allowed size (with a small headroom).
        keep = max_bytes // 2
        trimmed = data[-keep:]
        # Start on a newline boundary when possible.
        nl = trimmed.find(b"\n")
        if 0 <= nl < len(trimmed) - 1:
            trimmed = trimmed[nl + 1 :]
        header = (
            f"===== log trimmed to ~{keep // (1024 * 1024)}MB "
            f"(limit {max_bytes // (1024 * 1024)}MB) =====\n"
        ).encode("utf-8")
        log_path.write_bytes(header + trimmed)
    except OSError:
        pass


def write_report(
    hits: list[Hit],
    reports_dir: Path,
    generated_at: datetime,
    duration_seconds: float,
    unavailable_sites: list[str] | None = None,
) -> Path:
    reports_dir.mkdir(parents=True, exist_ok=True)
    stamp = generated_at.strftime("%Y-%m-%d_%H-%M-%S")
    out_path = reports_dir / f"media-monitor-{stamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Hits"

    headers = (
        "слово / фраза",
        "ссылка на страницу",
        "дата выхода статьи",
        "название статьи",
        "дата время скана",
        "версия ПО",
        "длительность генерации",
    )
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")

    # F1 — версия ПО; G1 — длительность генерации отчёта
    ws["F1"] = f"v{VERSION}"
    ws["G1"] = format_duration(duration_seconds)
    ws["F1"].font = Font(bold=True)
    ws["G1"].font = Font(bold=True)

    for hit in hits:
        ws.append(
            (
                hit.phrase,
                hit.url,
                hit.published_at,
                hit.title,
                hit.scanned_at.strftime("%Y-%m-%d %H:%M:%S"),
                "",
                "",
            )
        )

    hit_last_row = ws.max_row
    if hit_last_row >= 2:
        for row in ws.iter_rows(min_row=2, max_col=2, max_row=hit_last_row):
            cell = row[1]
            if cell.value and str(cell.value).startswith("http"):
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"

    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    for site in unavailable_sites or []:
        ws.append((site, "недоступен", "", "", "", "", ""))
        row_idx = ws.max_row
        for col in (1, 2):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = red_fill
            cell.font = Font(bold=True, color="FFFFFF")
        if str(site).startswith("http"):
            ws.cell(row=row_idx, column=1).hyperlink = str(site)

    widths = (36, 70, 20, 50, 22, 14, 22)
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    wb.save(out_path)
    return out_path


def run(*, interactive_auth: bool = True) -> int:
    root = app_dir()
    reports_dir = root / "reports"
    started_mono = time.monotonic()
    clear_cancel()

    log_print(f"media-monitor v{VERSION}")
    with _SSL_LOCK:
        _INSECURE_SSL_HOSTS.clear()

    try:
        sites_path = ensure_local_config(root, "sites.txt")
        words_path = ensure_local_config(root, "words.txt")
        exclude_path = ensure_local_config(root, "exclude.txt")
        settings_path = resolve_settings_path(root)
        sync_settings_file(settings_path)
        sites = load_lines(sites_path)
        phrases = load_lines(words_path)
        excluded = load_exclude_urls(exclude_path)
        settings = load_settings(settings_path)
    except (FileNotFoundError, ValueError) as exc:
        log_print(f"ERROR: {exc}")
        elapsed = time.monotonic() - started_mono
        log_print(f"Elapsed: {format_duration(elapsed)} ({elapsed:.1f}s)")
        log_print(f"media-monitor v{VERSION}")
        return 1

    log_print(f"Папка: {root}")
    log_print(
        f"Конфиги: {sites_path.name}, {words_path.name}, "
        f"{settings_path.name}, {exclude_path.name}"
        + (f" (исключено URL: {len(excluded)})" if excluded else "")
    )
    older = settings.article_date_not_older_than
    later = settings.article_date_not_later_than
    eff_from, eff_to = effective_article_date_range(
        older, later, settings.article_date_last_days
    )
    range_ru = format_article_date_range_ru(eff_from, eff_to)
    if settings.article_date_last_days > 0:
        log_print(
            f"Период статей: {range_ru} "
            f"(последние {settings.article_date_last_days} дн.; "
            "статьи без даты тоже попадают в отчёт)"
        )
    else:
        log_print(
            f"Период статей: {range_ru} "
            "(статьи без даты тоже попадают в отчёт)"
        )
    scan_limit = settings.max_scan_urls
    log_print(
        "Лимиты: страниц на сайт = "
        f"{'без лимита' if scan_limit == 0 else scan_limit}, "
        f"ссылок со страницы = {settings.max_expand_links}"
    )
    log_print(
        "SSL: "
        + ("проверка вкл. (встроенные CA + Минцифры)" if settings.ssl_verify else "без проверки")
    )

    if not sites:
        log_print("ERROR: sites.txt is empty (no URLs).")
        elapsed = time.monotonic() - started_mono
        log_print(f"Elapsed: {format_duration(elapsed)} ({elapsed:.1f}s)")
        log_print(f"media-monitor v{VERSION}")
        return 1
    if not phrases:
        log_print("ERROR: words.txt is empty (no keywords/phrases).")
        elapsed = time.monotonic() - started_mono
        log_print(f"Elapsed: {format_duration(elapsed)} ({elapsed:.1f}s)")
        log_print(f"media-monitor v{VERSION}")
        return 1

    site_groups = group_seeds_by_origin(sites)
    total_sites = len(site_groups)
    workers = settings.site_workers if settings.site_workers > 0 else total_sites
    workers = max(1, min(workers, total_sites))

    log_print(
        f"К работе: {total_sites} сайт(ов), {len(phrases)} фраз(ы). "
        f"Параллельно сайтов: {workers} "
        f"(внутри каждого сайта запросы идут по одному)."
    )
    if settings.log_verbose:
        log_print("Подробный лог URL включён (log_verbose=1).")
    else:
        log_print(
            "Вкладка «Статус» — итог по сайтам. "
            "Вкладка «Полный лог» — что качается прямо сейчас. "
            "Сводка также в верхней панели."
        )

    if interactive_auth:
        auth = prompt_credentials(settings.auth_timeout_seconds)
    else:
        auth = None
        log_print("Авторизация: пропуск (режим окна).")

    scanned_at = datetime.now()
    global _CA_BUNDLE, _LOG_VERBOSE
    _LOG_VERBOSE = settings.log_verbose
    _CA_BUNDLE = build_ca_bundle() if settings.ssl_verify else False

    progress = RunProgress(total_sites=total_sites, workers=workers)
    progress._notify()
    hits: list[Hit] = []
    errors = 0
    skipped_auth = 0
    pages_total = 0
    fatal_sites = 0
    unavailable: list[str] = []
    unavailable_hosts: set[str] = set()
    cancelled = False
    stopped_sites = 0

    log_print("Сканирование…")
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {
            pool.submit(
                process_site,
                index,
                total_sites,
                origin,
                seeds,
                phrases,
                settings,
                auth,
                excluded,
                scanned_at,
                progress,
            ): origin
            for index, (origin, seeds) in enumerate(site_groups, start=1)
        }
        for fut in as_completed(futures):
            # Soft-cancel: mark flag for workers, cancel queued tasks,
            # but ALWAYS collect results so Excel keeps found hits.
            if is_cancelled():
                cancelled = True
                for pending in futures:
                    pending.cancel()
            try:
                site_result = fut.result()
            except FuturesCancelledError:
                stopped_sites += 1
                with progress._lock:
                    progress.done += 1
                    progress.cancelled_sites += 1
                progress._notify()
                continue
            except Exception as exc:  # noqa: BLE001
                log_print(f"ERROR worker: {exc}")
                errors += 1
                continue
            hits.extend(site_result.hits)
            errors += site_result.errors
            skipped_auth += site_result.skipped_auth
            pages_total += site_result.pages_scanned
            if site_result.fatal == "cancelled":
                stopped_sites += 1
            elif site_result.fatal:
                fatal_sites += 1
                errors += 1
            if site_result.unavailable:
                unavailable.append(site_result.origin)
                unavailable_hosts.add(site_host_key(site_result.origin))

    if cancelled or is_cancelled():
        log_print()
        log_print(
            "Остановка завершена: сеть закрыта, оставшиеся страницы не сканировались."
        )
        if stopped_sites:
            log_print(f"Остановлено сайтов: {stopped_sites}.")
        log_print("Excel всё равно сохраняется с тем, что успели найти.")

    if unavailable and settings.comment_unavailable_sites:
        n = comment_unavailable_in_sites(sites_path, unavailable_hosts)
        if n:
            log_print(
                f"sites.txt: закомментировано недоступных строк: {n} "
                f"({', '.join(sorted(unavailable_hosts))})"
            )

    # Curator: one Excel write after all workers finish.
    report_path = write_report(
        hits,
        reports_dir,
        datetime.now(),
        time.monotonic() - started_mono,
        unavailable_sites=unavailable,
    )
    elapsed = time.monotonic() - started_mono
    log_print()
    log_print(
        f"Итог: сайтов {total_sites}, страниц {pages_total}, "
        f"находок {len(hits)}, ошибок {errors}"
        + (f", недоступных {len(unavailable)}" if unavailable else "")
        + (f", остановлено {stopped_sites}" if stopped_sites else "")
        + (f", пропуск auth {skipped_auth}" if skipped_auth else "")
        + ("." if not (cancelled or is_cancelled()) else " (запуск прерван).")
    )
    log_print(f"Отчёт: {report_path}")
    if try_open_file(report_path):
        log_print("Отчёт открыт в Excel (если установлено приложение по умолчанию).")
    else:
        log_print("Отчёт не удалось открыть автоматически — откройте папку «Отчёты».")
    log_print(f"Время: {format_duration(elapsed)} ({elapsed:.1f}s)")
    log_print(f"media-monitor v{VERSION}")
    return 0


def wait_for_enter() -> None:
    try:
        input("\nНажмите Enter для выхода...")
    except EOFError:
        pass


def console_main() -> int:
    """CLI mode (also: media-monitor-vX.exe --console)."""
    if getattr(sys, "frozen", False) and sys.platform == "win32":
        try:
            import ctypes

            ctypes.windll.kernel32.AllocConsole()
            sys.stdout = open("CONOUT$", "w", encoding="utf-8", errors="replace")
            sys.stderr = open("CONOUT$", "w", encoding="utf-8", errors="replace")
            sys.stdin = open("CONIN$", "r", encoding="utf-8", errors="replace")
        except Exception:
            pass

    root = app_dir()
    log_path = root / LOG_NAME
    rotate_log_if_needed(log_path)
    original_stdout = sys.stdout
    original_stderr = sys.stderr
    log_file = open(log_path, "a", encoding="utf-8")
    code = 1
    try:
        started = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_file.write(f"\n===== media-monitor v{VERSION} start {started} =====\n")
        log_file.flush()
        sys.stdout = Tee(original_stdout, log_file)  # type: ignore[assignment]
        sys.stderr = Tee(original_stderr, log_file)  # type: ignore[assignment]
        code = run(interactive_auth=True)
        return code
    finally:
        sys.stdout = original_stdout
        sys.stderr = original_stderr
        ended = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_file.write(f"===== media-monitor v{VERSION} end {ended} =====\n")
        log_file.close()
        rotate_log_if_needed(log_path)
        if getattr(sys, "frozen", False):
            wait_for_enter()


def main() -> int:
    ensure_stdio()
    if "--console" in sys.argv:
        return console_main()
    from gui import run_gui

    return run_gui()


if __name__ == "__main__":
    raise SystemExit(main())
